from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import render_to_string
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Count, Q
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from io import BytesIO

from carreras.models import Carrera, PlanEstudio
from programas.models import Programa
from universidades.models import Universidad, Sede, Facultad
from seguimiento.models import ProcesoCurricular, FaseProceso, TipoFase
from formularios.models import FormularioValoracion


# ── Helpers ───────────────────────────────────────────────────────────────────

def _stats_base():
    """Calcula las estadísticas base usadas en varios endpoints."""
    from seguimiento.models import ProcesoCurricular as PC
    carreras  = Carrera.objects.prefetch_related('procesos').filter(
                    en_funcionamiento=True)
    programas = Programa.objects.filter(activo=True)
    procesos  = ProcesoCurricular.objects.all()

    sedes_unicas = Sede.objects.filter(activa=True).values(
        'ciudad', 'facultad__universidad'
    ).distinct().count()

    todos_ids    = set(carreras.values_list('id', flat=True))
    ids_rediseno = set(PC.objects.filter(tipo_proceso='REDISENO'
                       ).values_list('carrera_id', flat=True))
    ids_diseno   = set(PC.objects.filter(tipo_proceso='DISENO'
                       ).values_list('carrera_id', flat=True))
    ids_ajuste   = set(PC.objects.filter(tipo_proceso='AJUSTE'
                       ).values_list('carrera_id', flat=True))
    ids_compl    = set(PC.objects.filter(tipo_proceso='COMPLEMENTACION'
                       ).values_list('carrera_id', flat=True))
    ids_con      = ids_rediseno | ids_diseno | ids_ajuste | ids_compl
    total_int    = len(todos_ids) or 1

    return {
        'carreras':   carreras,
        'programas':  programas,
        'procesos':   procesos,
        'stats': {
            'total_carreras':       carreras.count(),
            'total_programas':      programas.count(),
            'total_universidades':  Universidad.objects.filter(activa=True).count(),
            'total_sedes':          sedes_unicas,
            'total_procesos':       procesos.count(),
            'procesos_en_proceso':  procesos.filter(estado='EN_PROCESO').count(),
            'procesos_concluidos':  procesos.filter(estado='CONCLUIDO').count(),
            'formularios_total':    FormularioValoracion.objects.count(),
            'formularios_aprobados': FormularioValoracion.objects.filter(
                                         estado='APROBADO').count(),
            'vigentes':  sum(1 for c in carreras if c.estado_rediseno == 'VIGENTE'),
            'proximas':  sum(1 for c in carreras if c.estado_rediseno == 'PROXIMO'),
            'vencidas':  sum(1 for c in carreras if c.estado_rediseno == 'VENCIDO'),
            'sin_datos': sum(1 for c in carreras if c.estado_rediseno == 'SIN_DATOS'),
            'con_rediseno':        len(ids_rediseno & todos_ids),
            'con_diseno':          len(ids_diseno & todos_ids),
            'con_ajuste':          len(ids_ajuste & todos_ids),
            'con_complementacion': len(ids_compl & todos_ids),
            'sin_proceso':         len(todos_ids - ids_con),
            'total_carreras_int':  len(todos_ids),
            'pct_rediseno':        round(len(ids_rediseno & todos_ids) / total_int * 100),
            'pct_diseno':          round(len(ids_diseno & todos_ids)   / total_int * 100),
            'pct_ajuste':          round(len(ids_ajuste & todos_ids)   / total_int * 100),
            'pct_complementacion': round(len(ids_compl & todos_ids)    / total_int * 100),
            'pct_sin_proceso':     round(len(todos_ids - ids_con)      / total_int * 100),
        }
    }


# ── Vista principal de estadísticas ──────────────────────────────────────────

@login_required
def vista_estadisticas(request):
    from django.db.models import Count

    data = _stats_base()

    sedes_por_depto = list(
        Sede.objects.filter(activa=True).values(
            'departamento', 'ciudad', 'facultad__universidad'
        ).distinct().values('departamento').annotate(
            total=Count('ciudad', distinct=True)
        ).order_by('departamento')
    )

    carreras_por_universidad = list(
        Carrera.objects.filter(en_funcionamiento=True).values(
            'sede__facultad__universidad__sigla'
        ).annotate(total=Count('id')).order_by('-total')[:10]
    )

    programas_por_estado = {
        'FORMULACION':    data['programas'].filter(estado='FORMULACION').count(),
        'APROBACION':     data['programas'].filter(estado='APROBACION').count(),
        'IMPLEMENTACION': data['programas'].filter(estado='IMPLEMENTACION').count(),
        'SUSPENDIDO':     data['programas'].filter(estado='SUSPENDIDO').count(),
    }

    return render(request, 'reportes/estadisticas.html', {
        'stats':                    data['stats'],
        'sedes_por_depto':          sedes_por_depto,
        'carreras_por_universidad': carreras_por_universidad,
        'programas_por_estado':     programas_por_estado,
    })


# ── Buscador por carrera específica ──────────────────────────────────────────

@login_required
def buscador_carreras(request):
    """
    Buscador con filtros: Universidad + Carrera + Estado de proceso.
    Muestra el detalle de cada carrera: fases completadas, fase actual,
    año del último rediseño, etc.
    """
    # Parámetros de búsqueda
    q_universidad = request.GET.get('universidad', '')
    q_carrera     = request.GET.get('carrera', '')
    q_estado      = request.GET.get('estado_proceso', '')

    carreras_qs = Carrera.objects.select_related(
        'sede__facultad__universidad'
    ).prefetch_related(
        'procesos__fases__tipo_fase',
        'planes_estudio'
    ).filter(en_funcionamiento=True)

    if q_universidad:
        carreras_qs = carreras_qs.filter(
            sede__facultad__universidad__id=q_universidad
        )
    if q_carrera:
        carreras_qs = carreras_qs.filter(
            Q(nombre__icontains=q_carrera) |
            Q(titulo_profesional__icontains=q_carrera)
        )
    if q_estado:
        if q_estado == 'SIN_PROCESO':
            carreras_qs = carreras_qs.filter(procesos__isnull=True)
        else:
            carreras_qs = carreras_qs.filter(
                procesos__tipo_proceso=q_estado
            ).distinct()

    # Enriquecer cada carrera con datos del proceso
    resultados = []
    for carrera in carreras_qs.order_by(
            'sede__facultad__universidad__sigla', 'nombre'):
        ultimo_proceso = carrera.procesos.order_by('-anio_inicio').first()
        fase_actual    = None
        fases_ok       = 0
        total_fases    = 0

        if ultimo_proceso:
            fases       = list(ultimo_proceso.fases.all().order_by(
                              'tipo_fase__orden'))
            total_fases = len(fases)
            fases_ok    = sum(1 for f in fases if f.estado == 'COMPLETADO')
            # Fase actual = primera EN_PROCESO, si no hay, primera PENDIENTE
            fase_actual = next(
                (f for f in fases if f.estado == 'EN_PROCESO'), None
            ) or next(
                (f for f in fases if f.estado == 'PENDIENTE'), None
            )

        ultimo_plan = carrera.planes_estudio.order_by(
            '-anio_aprobacion').first()

        resultados.append({
            'carrera':          carrera,
            'ultimo_proceso':   ultimo_proceso,
            'fase_actual':      fase_actual,
            'fases_ok':         fases_ok,
            'total_fases':      total_fases,
            'ultimo_plan':      ultimo_plan,
            'estado_rediseno':  carrera.estado_rediseno,
        })

    universidades = Universidad.objects.filter(activa=True).order_by('sigla')

    TIPOS_PROCESO = [
        ('DISENO',         'Diseño Curricular'),
        ('REDISENO',       'Rediseño Curricular'),
        ('AJUSTE',         'Ajuste Curricular'),
        ('COMPLEMENTACION','Complementación Curricular'),
        ('SIN_PROCESO',    'Sin proceso registrado'),
    ]

    return render(request, 'reportes/buscador_carreras.html', {
        'resultados':      resultados,
        'universidades':   universidades,
        'tipos_proceso':   TIPOS_PROCESO,
        'q_universidad':   q_universidad,
        'q_carrera':       q_carrera,
        'q_estado':        q_estado,
        'total':           len(resultados),
        'con_filtro':      bool(q_universidad or q_carrera or q_estado),
    })


# ── Excel detallado (3 hojas) ─────────────────────────────────────────────────

@login_required
def exportar_carreras_excel(request):
    from django.utils import timezone

    wb = openpyxl.Workbook()

    # Estilos reutilizables
    COLOR_CABECERA   = '0F1F3D'
    COLOR_SUBHEADER  = '1A3560'
    COLOR_FILA_PAR   = 'EEF2F8'
    COLOR_ACCENT     = 'E8A020'

    estilo_titulo = Font(bold=True, color='FFFFFF', size=11)
    estilo_sub    = Font(bold=True, color='FFFFFF', size=10)
    fill_header   = PatternFill('solid', fgColor=COLOR_CABECERA)
    fill_sub      = PatternFill('solid', fgColor=COLOR_SUBHEADER)
    fill_par      = PatternFill('solid', fgColor=COLOR_FILA_PAR)
    centro        = Alignment(horizontal='center', vertical='center',
                               wrap_text=True)
    izquierda     = Alignment(horizontal='left',   vertical='center',
                               wrap_text=True)
    borde_fino    = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def aplicar_cabecera(ws, fila, cols, fill=fill_header, font=estilo_titulo):
        for col, texto in enumerate(cols, 1):
            c = ws.cell(row=fila, column=col, value=texto)
            c.fill      = fill
            c.font      = font
            c.alignment = centro
            c.border    = borde_fino

    def fila_datos(ws, fila, valores, par=False):
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=fila, column=col, value=val)
            c.alignment = izquierda
            c.border    = borde_fino
            if par:
                c.fill = fill_par

    # ════════════════════════════════════════════════════
    # HOJA 1 — Resumen General
    # ════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = '1. Resumen General'
    ws1.sheet_view.showGridLines = False

    data = _stats_base()
    stats = data['stats']

    # Título principal
    ws1.merge_cells('A1:D1')
    c = ws1['A1']
    c.value     = 'SIGCU Bolivia — Resumen Estadístico General'
    c.font      = Font(bold=True, size=14, color=COLOR_CABECERA)
    c.alignment = centro
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells('A2:D2')
    c = ws1['A2']
    from django.utils import timezone as tz
    c.value     = (f'Universidad Autónoma Tomás Frías — Potosí, Bolivia   '
                   f'| Generado: {tz.now().strftime("%d/%m/%Y %H:%M")}')
    c.font      = Font(italic=True, size=9, color='666666')
    c.alignment = centro
    ws1.row_dimensions[2].height = 16

    # Separador
    ws1.row_dimensions[3].height = 8

    # Bloque de totales
    titulos_totales = [
        ('Universidades',    stats['total_universidades']),
        ('Sedes activas',    stats['total_sedes']),
        ('Carreras',         stats['total_carreras']),
        ('Programas',        stats['total_programas']),
        ('Procesos',         stats['total_procesos']),
        ('Formularios',      stats['formularios_total']),
    ]
    ws1.cell(row=4, column=1, value='INDICADOR').font = Font(bold=True)
    ws1.cell(row=4, column=2, value='CANTIDAD').font  = Font(bold=True)
    aplicar_cabecera(ws1, 4, ['INDICADOR', 'CANTIDAD', '', ''])
    for i, (label, val) in enumerate(titulos_totales, 5):
        fila_datos(ws1, i, [label, val, '', ''], par=(i % 2 == 0))

    ws1.row_dimensions[11].height = 10

    # Estado rediseño
    aplicar_cabecera(ws1, 12,
        ['ESTADO REDISEÑO', 'CANTIDAD', 'PORCENTAJE', ''],
        fill=fill_sub, font=estilo_sub)
    total = stats['total_carreras_int'] or 1
    datos_rediseno = [
        ('Vigentes',           stats['vigentes'],
         f"{round(stats['vigentes']/total*100)}%"),
        ('Próximos a vencer',  stats['proximas'],
         f"{round(stats['proximas']/total*100)}%"),
        ('Vencidos',           stats['vencidas'],
         f"{round(stats['vencidas']/total*100)}%"),
        ('Sin datos',          stats['sin_datos'],
         f"{round(stats['sin_datos']/total*100)}%"),
    ]
    for i, (label, val, pct) in enumerate(datos_rediseno, 13):
        fila_datos(ws1, i, [label, val, pct, ''], par=(i % 2 == 0))

    ws1.row_dimensions[17].height = 10

    # Procesos por tipo
    aplicar_cabecera(ws1, 18,
        ['TIPO DE INNOVACIÓN CURRICULAR', 'CARRERAS', 'PORCENTAJE', ''],
        fill=fill_sub, font=estilo_sub)
    datos_proc = [
        ('Rediseño Curricular',        stats['con_rediseno'],
         f"{stats['pct_rediseno']}%"),
        ('Diseño Curricular',          stats['con_diseno'],
         f"{stats['pct_diseno']}%"),
        ('Ajuste Curricular',          stats['con_ajuste'],
         f"{stats['pct_ajuste']}%"),
        ('Complementación Curricular', stats['con_complementacion'],
         f"{stats['pct_complementacion']}%"),
        ('Sin proceso registrado',     stats['sin_proceso'],
         f"{stats['pct_sin_proceso']}%"),
    ]
    for i, (label, val, pct) in enumerate(datos_proc, 19):
        fila_datos(ws1, i, [label, val, pct, ''], par=(i % 2 == 0))

    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 10

    # ════════════════════════════════════════════════════
    # HOJA 2 — Detalle Carreras
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet('2. Detalle Carreras')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:N1')
    c = ws2['A1']
    c.value     = 'SIGCU Bolivia — Detalle Completo de Carreras'
    c.font      = Font(bold=True, size=13, color=COLOR_CABECERA)
    c.alignment = centro
    ws2.row_dimensions[1].height = 24

    cols2 = [
        'N°', 'Universidad', 'Facultad', 'Sede', 'Ciudad',
        'Departamento', 'Carrera', 'Grado', 'Título Profesional',
        'N° SUB', 'Área', 'Enfoque', 'Funcionando', 'Estado Rediseño'
    ]
    aplicar_cabecera(ws2, 2, cols2)
    ws2.row_dimensions[2].height = 32

    carreras = Carrera.objects.select_related(
        'sede__facultad__universidad'
    ).prefetch_related('procesos', 'planes_estudio').filter(
        en_funcionamiento=True
    ).order_by('sede__facultad__universidad__sigla',
               'sede__facultad__nombre', 'nombre')

    for i, c in enumerate(carreras, 1):
        par = i % 2 == 0
        fila_datos(ws2, i + 2, [
            i,
            c.sede.facultad.universidad.sigla,
            c.sede.facultad.nombre,
            c.sede.nombre,
            c.sede.ciudad,
            c.sede.departamento,
            c.nombre,
            c.get_grado_display(),
            c.titulo_profesional or '—',
            c.numero_sub or '—',
            c.area,
            c.get_enfoque_curricular_display() if c.enfoque_curricular else '—',
            'Sí' if c.en_funcionamiento else 'No',
            c.estado_rediseno,
        ], par=par)

    anchos2 = [5,10,40,35,14,14,45,14,40,10,8,14,12,16]
    for col, ancho in enumerate(anchos2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = ancho

    ws2.freeze_panes = 'A3'

    # ════════════════════════════════════════════════════
    # HOJA 3 — Procesos y Fases
    # ════════════════════════════════════════════════════
    ws3 = wb.create_sheet('3. Procesos y Fases')
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:M1')
    c = ws3['A1']
    c.value     = 'SIGCU Bolivia — Seguimiento de Procesos Curriculares y Fases'
    c.font      = Font(bold=True, size=13, color=COLOR_CABECERA)
    c.alignment = centro
    ws3.row_dimensions[1].height = 24

    cols3 = [
        'N°', 'Universidad', 'Carrera', 'Ciudad', 'Tipo de Proceso',
        'Año Inicio', 'Año Conclusión', 'Estado Proceso',
        'Fases Completadas', 'Total Fases', '% Avance',
        'Fase Actual', 'Resolución HCU/RAN'
    ]
    aplicar_cabecera(ws3, 2, cols3)
    ws3.row_dimensions[2].height = 32

    procesos = ProcesoCurricular.objects.select_related(
        'carrera__sede__facultad__universidad'
    ).prefetch_related('fases__tipo_fase').order_by(
        'carrera__sede__facultad__universidad__sigla',
        'carrera__nombre', '-anio_inicio'
    )

    for i, p in enumerate(procesos, 1):
        par        = i % 2 == 0
        fases      = list(p.fases.all())
        total_f    = len(fases)
        completadas = sum(1 for f in fases if f.estado == 'COMPLETADO')
        pct        = round(completadas / total_f * 100) if total_f else 0
        fase_act   = next(
            (f.tipo_fase.nombre for f in fases if f.estado == 'EN_PROCESO'),
            next(
                (f.tipo_fase.nombre for f in fases if f.estado == 'PENDIENTE'),
                '—'
            )
        )
        plan = p.carrera.planes_estudio.order_by('-anio_aprobacion').first()
        resolucion = (f"HCU: {plan.resolucion_hcu} / RAN: {plan.resolucion_ran}"
                      if plan and (plan.resolucion_hcu or plan.resolucion_ran)
                      else '—')
        fila_datos(ws3, i + 2, [
            i,
            p.carrera.sede.facultad.universidad.sigla,
            p.carrera.nombre,
            p.carrera.sede.ciudad,
            p.get_tipo_proceso_display(),
            p.anio_inicio,
            p.anio_conclusion or '—',
            p.get_estado_display(),
            completadas,
            total_f,
            f"{pct}%",
            fase_act,
            resolucion,
        ], par=par)

    anchos3 = [5,12,42,14,22,10,14,14,16,12,10,30,30]
    for col, ancho in enumerate(anchos3, 1):
        ws3.column_dimensions[get_column_letter(col)].width = ancho

    ws3.freeze_panes = 'A3'

    # Guardar y devolver
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=sigcu_reporte_completo.xlsx'
    wb.save(response)
    return response


# ── Excel de búsqueda específica ─────────────────────────────────────────────

@login_required
def exportar_busqueda_excel(request):
    """Exporta los resultados del buscador a Excel."""
    from django.db.models import Q

    q_universidad = request.GET.get('universidad', '')
    q_carrera     = request.GET.get('carrera', '')
    q_estado      = request.GET.get('estado_proceso', '')

    carreras_qs = Carrera.objects.select_related(
        'sede__facultad__universidad'
    ).prefetch_related(
        'procesos__fases__tipo_fase', 'planes_estudio'
    ).filter(en_funcionamiento=True)

    if q_universidad:
        carreras_qs = carreras_qs.filter(
            sede__facultad__universidad__id=q_universidad)
    if q_carrera:
        carreras_qs = carreras_qs.filter(
            Q(nombre__icontains=q_carrera) |
            Q(titulo_profesional__icontains=q_carrera))
    if q_estado == 'SIN_PROCESO':
        carreras_qs = carreras_qs.filter(procesos__isnull=True)
    elif q_estado:
        carreras_qs = carreras_qs.filter(
            procesos__tipo_proceso=q_estado).distinct()

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Búsqueda SIGCU'
    ws.sheet_view.showGridLines = False

    fill_h = PatternFill('solid', fgColor='0F1F3D')
    font_h = Font(bold=True, color='FFFFFF', size=10)
    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    izq    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    borde  = Border(
        left=Side(style='thin',   color='CCCCCC'),
        right=Side(style='thin',  color='CCCCCC'),
        top=Side(style='thin',    color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.merge_cells('A1:L1')
    ws['A1'].value     = 'SIGCU Bolivia — Resultado de Búsqueda de Carreras'
    ws['A1'].font      = Font(bold=True, size=12, color='0F1F3D')
    ws['A1'].alignment = centro
    ws.row_dimensions[1].height = 22

    cols = ['N°','Universidad','Facultad','Carrera','Grado',
            'Título Profesional','N° SUB','Ciudad',
            'Tipo de Proceso','Estado Proceso',
            'Fase Actual','Estado Rediseño']
    for col, h in enumerate(cols, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = fill_h; c.font = font_h
        c.alignment = centro; c.border = borde
    ws.row_dimensions[2].height = 28

    fill_par = PatternFill('solid', fgColor='EEF2F8')
    for i, carrera in enumerate(
            carreras_qs.order_by(
                'sede__facultad__universidad__sigla', 'nombre'), 1):
        ultimo = carrera.procesos.order_by('-anio_inicio').first()
        tipo_proc  = ultimo.get_tipo_proceso_display() if ultimo else '—'
        estado_proc = ultimo.get_estado_display()      if ultimo else '—'
        fase_act   = '—'
        if ultimo:
            fases = list(ultimo.fases.all())
            fa = next((f for f in fases if f.estado == 'EN_PROCESO'), None) \
              or next((f for f in fases if f.estado == 'PENDIENTE'), None)
            if fa:
                fase_act = f"{fa.tipo_fase.codigo} — {fa.tipo_fase.nombre}"

        row = i + 2
        par = i % 2 == 0
        valores = [
            i,
            carrera.sede.facultad.universidad.sigla,
            carrera.sede.facultad.nombre,
            carrera.nombre,
            carrera.get_grado_display(),
            carrera.titulo_profesional or '—',
            carrera.numero_sub or '—',
            carrera.sede.ciudad,
            tipo_proc, estado_proc, fase_act,
            carrera.estado_rediseno,
        ]
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = izq; c.border = borde
            if par: c.fill = fill_par

    anchos = [5,12,38,42,14,38,10,12,22,16,35,16]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.freeze_panes = 'A3'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=busqueda_carreras.xlsx'
    wb.save(response)
    return response


# ── PDF de búsqueda específica ────────────────────────────────────────────────

@login_required
def exportar_busqueda_pdf(request):
    from xhtml2pdf import pisa
    from django.utils import timezone
    from django.db.models import Q

    q_universidad = request.GET.get('universidad', '')
    q_carrera     = request.GET.get('carrera', '')
    q_estado      = request.GET.get('estado_proceso', '')

    carreras_qs = Carrera.objects.select_related(
        'sede__facultad__universidad'
    ).prefetch_related(
        'procesos__fases__tipo_fase', 'planes_estudio'
    ).filter(en_funcionamiento=True)

    if q_universidad:
        carreras_qs = carreras_qs.filter(
            sede__facultad__universidad__id=q_universidad)
    if q_carrera:
        carreras_qs = carreras_qs.filter(
            Q(nombre__icontains=q_carrera) |
            Q(titulo_profesional__icontains=q_carrera))
    if q_estado == 'SIN_PROCESO':
        carreras_qs = carreras_qs.filter(procesos__isnull=True)
    elif q_estado:
        carreras_qs = carreras_qs.filter(
            procesos__tipo_proceso=q_estado).distinct()

    resultados = []
    for carrera in carreras_qs.order_by(
            'sede__facultad__universidad__sigla', 'nombre'):
        ultimo = carrera.procesos.order_by('-anio_inicio').first()
        fases_ok    = 0
        total_fases = 0
        fase_act    = None
        if ultimo:
            fases       = list(ultimo.fases.all())
            total_fases = len(fases)
            fases_ok    = sum(1 for f in fases if f.estado == 'COMPLETADO')
            fase_act    = next(
                (f for f in fases if f.estado == 'EN_PROCESO'), None
            ) or next((f for f in fases if f.estado == 'PENDIENTE'), None)
        resultados.append({
            'carrera':        carrera,
            'ultimo_proceso': ultimo,
            'fases_ok':       fases_ok,
            'total_fases':    total_fases,
            'fase_actual':    fase_act,
        })

    nombre_uni = ''
    if q_universidad:
        u = Universidad.objects.filter(id=q_universidad).first()
        if u: nombre_uni = u.nombre

    html_string = render_to_string('reportes/busqueda_pdf.html', {
        'resultados':   resultados,
        'fecha':        timezone.now(),
        'filtros': {
            'universidad': nombre_uni,
            'carrera':     q_carrera,
            'estado':      q_estado,
        },
        'total': len(resultados),
    })
    buffer = BytesIO()
    pisa.CreatePDF(html_string, dest=buffer, encoding='utf-8')
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=busqueda_carreras.pdf'
    return response


# ── PDF estadísticas generales mejorado ──────────────────────────────────────

@login_required
def exportar_estadisticas_pdf(request):
    from xhtml2pdf import pisa
    from django.utils import timezone
    from django.db.models import Count

    data  = _stats_base()
    stats = data['stats']

    sedes_por_depto = list(
        Sede.objects.filter(activa=True).values(
            'departamento', 'ciudad', 'facultad__universidad'
        ).distinct().values('departamento').annotate(
            total=Count('ciudad', distinct=True)
        ).order_by('departamento')
    )

    carreras_por_universidad = list(
        Carrera.objects.filter(en_funcionamiento=True).values(
            'sede__facultad__universidad__sigla',
            'sede__facultad__universidad__nombre',
        ).annotate(total=Count('id')).order_by('-total')
    )

    html_string = render_to_string('reportes/estadisticas_pdf.html', {
        'stats':                    stats,
        'sedes_por_depto':          sedes_por_depto,
        'carreras_por_universidad': carreras_por_universidad,
        'fecha':                    timezone.now(),
    })
    buffer = BytesIO()
    pisa.CreatePDF(html_string, dest=buffer, encoding='utf-8')
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=estadisticas_sigcu.pdf'
    return response


# ── Endpoint JSON ─────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def resumen_estadisticas(request):
    data  = _stats_base()
    return Response(data['stats'])